from django.db import models
from django.utils.timezone import now
from django.utils.text import slugify
from django.core.exceptions import ValidationError
from django.apps import apps
from project import settings
from apps.shop.models import *
from apps.productloader.tasks import load_products_worker, load_products_images_worker
from PIL import Image
from unidecode import unidecode
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import datetime
import time
import os
import rarfile
import shutil







# PRODUCTS LOADER
def validate_xlsx(value):
    ext = os.path.splitext(value.name)[1]  # [0] returns path+filename
    valid_extensions = ['.xlsx']
    if not ext.lower() in valid_extensions:
        raise ValidationError(u'Загрузите таблицу EXCEL в формате .xlsx')

def validate_rar(value):
    ext = os.path.splitext(value.name)[1]  # [0] returns path+filename
    valid_extensions = ['.rar']
    if not ext.lower() in valid_extensions:
        raise ValidationError(u'Загрузите архив в формате .rar')

def FilePath(instance, file):
    filename = slugify(unidecode(instance.brand_name))
    ext = file.split('.')[-1]
    path = 'table_load/' + ext + '/' + filename + '.' + ext
    return path


class LoadProductsImages(models.Model):
    brand_name =     models.CharField(max_length=255, blank=False, verbose_name='Навзвание бренда', default='')
    date =           models.DateTimeField(default=now, verbose_name='Дата загрузки')
    # File
    rar =            models.FileField(upload_to=FilePath, validators=[validate_rar],  verbose_name='RAR архив с картинками товаров')
    path =           models.CharField(blank=True, null=True, max_length=2500)
    # Images
    log_rar =        models.TextField(default='', blank=True, verbose_name='Ошибка разархивации')
    log_skiped =     models.TextField(default='', blank=True, verbose_name='Уже существуют')
    log_loaded =     models.TextField(default='', blank=True, verbose_name='Загружено')
    log_error =      models.TextField(default='', blank=True, verbose_name='Ошибка файла')


    class Meta:
        ordering = ('-date',)
        verbose_name = 'Загрузка изображений'
        verbose_name_plural = 'Загрузка изображений'


    def extract_rar(self):
        self.log_rar = ''
        name = self.rar.name.split('/')[-1].split('.')[0]
        self.path = 'table_load/rar/' + name
        path = settings.MEDIA_ROOT + self.path
        if not os.path.exists(path):
            os.mkdir(path)
        rf = rarfile.RarFile(path + '.rar')
        for f in rf.infolist():
            fp = '/'.join([path, f.filename]) 
            if os.path.exists(fp) == False:
                try:
                    rf.extract(f.filename, path=path)
                except:
                    self.log_rar += fp + '\n'
        return path


    def write_images(self):
        self.log_loaded = ''
        self.log_skiped = ''
        for root, dirs, files in os.walk(settings.MEDIA_ROOT + self.path):
            for f in files:
                image_code = str(f).split('.')[0]
                file_path =  settings.MEDIA_ROOT + self.path + '/' + f
                variant = Variant.objects.filter(code=image_code).first()
       
                if variant:
                    if VariantImages.objects.filter(parent=variant).count() == 0:
                        try:
                            image = VariantImages(parent=variant, image=file_path)
                            image.save()
                            if image.image_thmb == {}:
                                self.log_error += image_code + '\n'
                            else:
                                self.log_loaded += image_code + '\n'
                        except:
                            self.log_error += image_code + '\n'
                    else:
                        self.log_skiped += image_code + '\n'

    def run_worker(self):
        self.extract_rar()
        self.write_images()
        super(LoadProductsImages, self).save()


    def save(self):
        super(LoadProductsImages, self).save()
        load_products_images_worker.delay(self.id)
        

    def delete(self):
        path = settings.MEDIA_ROOT + self.path
        if os.path.exists(path):
            shutil.rmtree(path)
        super(LoadProductsImages, self).delete()
        



class LoadProductsFromTable(models.Model):
    brand_name =     models.CharField(max_length=255, blank=False, verbose_name='Навзвание бренда', default='')
    date =           models.DateTimeField(default=now, verbose_name='Дата загрузки')
    # Table
    table =          models.FileField(upload_to=FilePath, validators=[validate_xlsx], verbose_name='XLSX таблица товаров')
    table_old_path = models.CharField(blank=True, null=True, editable=False, max_length=2500)
    
    

    class Meta:
        ordering = ('-date',)
        verbose_name = 'Загрузка товаров из таблицы'
        verbose_name_plural = 'Загрузка товаров из таблицы'

    def __str__(self):
        return ' '.join([str(self.pk), str(self.date), self.brand_name])


    def open_table(self):
        xlsx_table = load_workbook(self.table)
        self.table_old_path = self.table
        sheet = xlsx_table[xlsx_table.sheetnames[0]]
        titles = [col[0].value for col in sheet.iter_cols()]
        return sheet, titles


    def write_brand(self, row, titles):
        brand_name = row[titles.index('Бренд')]
        brand = Brand.objects.filter(name__iexact=brand_name).first()
        if not brand:
            brand = Brand(name=brand_name)
            brand.save()
        
        serie_name = row[titles.index('Серия бренда')]
        serie = BrandSeries.objects.filter(parent=brand, name__iexact=serie_name).first()
        if not serie:
            serie = BrandSeries(parent=brand, name=serie_name)
            serie.save()
        return brand, serie


    def write_categories(self, row, titles):
        categories = []
        taxonomy = GoogleTaxonomyCategories.objects.filter(id=473).first()
        parent = None
        for n in range(1,4):
            params = {
                'parent' : parent,
                'name' : row[titles.index(f'Категория уровень {n}')],
                'human' : row[titles.index(f'Название категории для каталога {n}')], 
            }
            if params.get('name') not in [None, '']: 
                category = Categories.objects.filter(parent=parent, slug=slugify(unidecode(params['name']))).first()
                if not category:
                    category = Categories(**params, taxonomy=taxonomy)
                    category.save()

                parent = category
                categories.append(category)
            else:
                break
        return categories

    def get_unit(self, row, titles):
        try: unit = Unit.objects.get(
            name=row[titles.index('Тип варината')]
        )
        except: unit = Unit(
            name=row[titles.index('Тип варината')], 
            unit=row[titles.index('Еденица измерения')]
        )
        return unit


    def wrtite_product(self, row, titles, category, brand, serie):
        # Product
        product_params = {
            "category" :    category,
            "brand" :       brand,
            "serie" :       serie,
            "name" :        row[titles.index('Оригинальное название')],
            "human" :       row[titles.index('Человеческое неазвание')],
            "description" : row[titles.index('Описание')],
            "composition" : row[titles.index('Состав')],
            "application" : row[titles.index('Применение')],
            "unit" : self.get_unit(row, titles),
        }
        try:
            product = Product.objects.get(name=product_params["name"], brand=brand)
        except:
            product = Product(**product_params)
            product.save()
       
        # Variasnt
        variant_params = {
            "parent" :         product,
            "value" :          row[titles.index('Значение варианта')],
            "code" :           row[titles.index('Артикул')], 
            "barcode" :        row[titles.index('Штрихкод')], 
            "price" :          row[titles.index('Цена розничная')], 
            "whoosale_price" : row[titles.index('Цена оптвая')], 
        }

        try:
            variant = Variant.objects.get(parent=product, code=variant_params.get('code'))
        except:
            variant = Variant()
        
        # Set new params
        for k, v in variant_params.items():
            setattr(variant, k, v)
    
        variant.save()
        return product, variant


              
    def iter_table(self, sheet, titles):
        for n, row in enumerate(sheet.values):
            print(n)

            if n == 0: continue
            if row[0] == None: continue
            # Brand
            brand, serie = self.write_brand(row, titles)
            # Categories
            
            categories = self.write_categories(row, titles)
            category = categories[-1]
            
            # Product
            product, variant = self.wrtite_product(row, titles, category, brand, serie)
            # Write images
           
            super(LoadProductsFromTable, self).save()


    def run_worker(self):
        sheet, titles = self.open_table()
        self.iter_table(sheet, titles)
        
    

    def save(self, *args, **kwargs):
        super(LoadProductsFromTable, self).save()
        load_products_worker.delay(self.id)
    